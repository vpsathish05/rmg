"""
Rate Card Parser — Unified rate lookup from both pricing XLSX files.

Sources:
  1. docs/pricing/2511_JMAN Pricing Tool (aligned with new JIN).xlsx
     - Sheet "Rate Card & Lookups": role × location × GBP/USD/EUR billing + cost rates
  2. docs/pricing/Cluster_Revenue_COE_Forecast.xlsx
     - Sheet "Rate_Card": role × location × USD day rate (simplified)

Output: unified dict keyed by (role, location) → RateEntry
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


# ── Data classes ────────────────────────────────────────────────────────────

@dataclass
class RateEntry:
    """Single rate card entry for a role at a location."""
    role: str
    location: str  # UK, IN, USA
    day_rate_usd: float
    day_rate_gbp: float = 0.0
    day_rate_eur: float = 0.0
    cost_gbp: float = 0.0
    cost_usd: float = 0.0
    cost_eur: float = 0.0
    margin_pct: float = 0.0  # (billing - cost) / billing

    def monthly_revenue(self, fte: float = 1.0, working_days: int = 22) -> float:
        """Estimate monthly revenue in USD for given FTE."""
        return self.day_rate_usd * working_days * fte


# ── Constants ───────────────────────────────────────────────────────────────

_DOCS_DIR = Path(__file__).resolve().parent.parent.parent / "docs" / "pricing"

PRICING_TOOL_FILE = "2511_JMAN Pricing Tool (aligned with new JIN).xlsx"
CLUSTER_FORECAST_FILE = "Cluster_Revenue_COE_Forecast.xlsx"

# FX rates from pricing tool
FX_GBP_TO_USD = 1.3
FX_GBP_TO_EUR = 1.2


# ── Parsers ─────────────────────────────────────────────────────────────────

def _parse_pricing_tool(filepath: Path) -> dict[tuple[str, str], RateEntry]:
    """
    Parse '2511_JMAN Pricing Tool' → Rate Card & Lookups sheet.
    
    Layout (row 9 is header):
      Col B: Role
      Col C: Employee Location
      Col D: Billing GBP
      Col E: Billing USD
      Col F: Billing EUR
      Col G: Cost GBP
      Col H: Cost USD
      Col I: Cost EUR
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb["Rate Card & Lookups"]

    rates: dict[tuple[str, str], RateEntry] = {}

    for row in ws.iter_rows(min_row=10, max_row=42, values_only=True):
        # Col indices: 0=composite_key, 1=role, 2=location, 3=bill_gbp, 4=bill_usd,
        # 5=bill_eur, 6=cost_gbp, 7=cost_usd, 8=cost_eur
        role = row[1]
        location = row[2]
        if not role or not location:
            continue

        role = str(role).strip()
        location = str(location).strip()

        # Normalize location
        loc_norm = _normalize_location(location)
        if not loc_norm:
            continue

        bill_gbp = _safe_float(row[3])
        bill_usd = _safe_float(row[4])
        bill_eur = _safe_float(row[5])
        cost_gbp = _safe_float(row[6])
        cost_usd = _safe_float(row[7])
        cost_eur = _safe_float(row[8])

        margin = 0.0
        if bill_gbp > 0:
            margin = round((bill_gbp - cost_gbp) / bill_gbp, 4)

        key = (role, loc_norm)
        rates[key] = RateEntry(
            role=role,
            location=loc_norm,
            day_rate_usd=bill_usd,
            day_rate_gbp=bill_gbp,
            day_rate_eur=bill_eur,
            cost_gbp=cost_gbp,
            cost_usd=cost_usd,
            cost_eur=cost_eur,
            margin_pct=margin,
        )

    wb.close()
    return rates


def _parse_cluster_forecast_rates(filepath: Path) -> dict[tuple[str, str], RateEntry]:
    """
    Parse 'Cluster_Revenue_COE_Forecast.xlsx' → Rate_Card sheet.
    
    Simpler format:
      Col A (row 4+): Role
      Col B: Location
      Col C: Day Rate (USD)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb["Rate_Card"]

    rates: dict[tuple[str, str], RateEntry] = {}

    for row in ws.iter_rows(min_row=4, max_row=35, values_only=True):
        role = row[0]
        location = row[1]
        rate_usd = row[2]

        if not role or not location or not rate_usd:
            continue

        role = str(role).strip()
        location = str(location).strip()
        loc_norm = _normalize_location(location)
        if not loc_norm:
            continue

        rate_usd_val = _safe_float(rate_usd)
        if rate_usd_val <= 0:
            continue

        # Convert USD to GBP estimate
        rate_gbp = round(rate_usd_val / FX_GBP_TO_USD, 2)

        key = (role, loc_norm)
        if key not in rates:  # Don't overwrite if pricing tool already has it
            rates[key] = RateEntry(
                role=role,
                location=loc_norm,
                day_rate_usd=rate_usd_val,
                day_rate_gbp=rate_gbp,
            )

    wb.close()
    return rates


# ── Unified Loader ──────────────────────────────────────────────────────────

_CACHE: Optional[dict[tuple[str, str], RateEntry]] = None


def load_rate_card(docs_dir: Optional[Path] = None) -> dict[tuple[str, str], RateEntry]:
    """
    Load and merge rate cards from both XLSX sources.
    Priority: Pricing Tool (more detail) > Cluster Forecast (simpler).
    Returns dict keyed by (role, location).
    """
    global _CACHE
    if _CACHE is not None:
        return _CACHE

    base = docs_dir or _DOCS_DIR

    # Start with cluster forecast rates (less detailed)
    cluster_path = base / CLUSTER_FORECAST_FILE
    if cluster_path.exists():
        rates = _parse_cluster_forecast_rates(cluster_path)
    else:
        rates = {}

    # Override/enrich with pricing tool (more detailed — has cost, margin)
    pricing_path = base / PRICING_TOOL_FILE
    if pricing_path.exists():
        pricing_rates = _parse_pricing_tool(pricing_path)
        rates.update(pricing_rates)  # pricing tool wins on conflicts

    _CACHE = rates
    return rates


def get_rate(role: str, location: str) -> Optional[RateEntry]:
    """Look up rate for a specific role and location."""
    rates = load_rate_card()
    key = (role, location)
    if key in rates:
        return rates[key]

    # Try fuzzy match — case-insensitive
    role_lower = role.lower().strip()
    loc_norm = _normalize_location(location) or location
    for (r, l), entry in rates.items():
        if r.lower() == role_lower and l == loc_norm:
            return entry

    return None


def get_all_rates() -> list[RateEntry]:
    """Return all rate entries as a list."""
    return list(load_rate_card().values())


def get_rates_by_location(location: str) -> list[RateEntry]:
    """Get all rates for a specific location."""
    loc_norm = _normalize_location(location) or location
    return [r for r in load_rate_card().values() if r.location == loc_norm]


def clear_cache():
    """Clear the cached rate card (for reloading)."""
    global _CACHE
    _CACHE = None


# ── Helpers ─────────────────────────────────────────────────────────────────

def _normalize_location(loc: str) -> Optional[str]:
    """Normalize location string to standard: UK, IN, USA."""
    loc_upper = loc.upper().strip()
    if loc_upper in ("UK", "LONDON", "UNITED KINGDOM"):
        return "UK"
    elif loc_upper in ("IN", "INDIA", "CHENNAI"):
        return "IN"
    elif loc_upper in ("USA", "US", "NEW YORK", "UNITED STATES"):
        return "USA"
    return None


def _safe_float(val) -> float:
    """Safely convert a value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ── CLI test ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    rates = load_rate_card()
    print(f"Loaded {len(rates)} rate entries\n")
    print(f"{'Role':<35} {'Loc':<5} {'USD/day':>10} {'GBP/day':>10} {'Margin':>8}")
    print("-" * 75)
    for (role, loc), entry in sorted(rates.items(), key=lambda x: (-x[1].day_rate_usd, x[0])):
        margin_str = f"{entry.margin_pct*100:.1f}%" if entry.margin_pct > 0 else "—"
        print(f"{role:<35} {loc:<5} {entry.day_rate_usd:>10,.0f} {entry.day_rate_gbp:>10,.0f} {margin_str:>8}")
